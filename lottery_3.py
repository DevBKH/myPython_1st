#1. 임의로 7개의 숫자 조합을 생성한다(1개는 행운번호)
#2. 사용자가 구매하는 갯수만큼 자동선택 번호를 생성한다. (6개)
#3. 위 두개의 숫자 조합을 비교하여 같은 갯수가 몇개인지 알려준다.
#4. 그리고 동일한 번호가 무엇이었는지도 보여준다.
#5. 자동생성된 번호들을 엑셀에 저장한다
#6. 당첨번호와 행운번호를 엑셀에 저장한다
#7. 최종으로 각 순위별 당첨갯수를 엑셀에 저장한다

from random import *
import openpyxl
myExcelPATH = "d:/siccBKH/Documents/Test/"
myExcelFile_1 = "myexcel.xlsx"

class Lotto :
    def __init__(self):
        self.winNum = self.__lotto_pop()
        self.winNum.sort()

    # 로또당첨번호 조합(7개 숫자)를 만드는 메소드(마지막 번호는 행운번호)
    @staticmethod
    def __lotto_pop():
        return sample(range(1,46),7)

    # 임의의 로또번호를 만드는 메소드(구매자가 구입하는 번호)
    def buyLotto(self) :
        myList = []
        chkVal = 0

        for i in range(0,6) :
            myList.append(randint(1,45))

        myList.sort()

        i = 0
        while i < len(myList) :
            if i > 0 and chkVal == myList[i] :
                myList.pop(i)
                myList.append(randint(1,45))
                myList.sort()
                i = 0

            chkVal = myList[i]
            i +=1
        return myList

    # 입력된 번호와 당첨번호를 비교하여 당첨번호와 순위를 리턴하는 메소드
    def compLottoNum(self, userNum) :
        self.num_match = list(set(userNum)&set(self.winNum[:6]))
        if len(self.num_match) == 6 :
            return 1
        elif len(self.num_match) == 5 and self.winNum[-1] in userNum :
            return 2
        elif len(self.num_match) >= 3 :
            return 8-len(self.num_match)
        else :
            return 0

    # 엑셀에 저장하는 메소드
    def saveLottoExcel(self, userNum, userRank) :
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LottoNum"

        for col in range(1,8) :
            ws.cell(row=1, column=1).value = "Win Num"
            ws.cell(row=1, column=col+1).value = self.winNum[col-1]

        for row in range(0,len(userNum)) :
            ws.cell(row=row+2, column=1).value = "Num"+"{}".format(row+1)
            for col in range(0,len(userNum[row])) :
                ws.cell(row=row+2, column=col+2).value = userNum[row][col]

        irow = len(userNum)+3
        if userRank == [0]*5 :
            ws.cell(row=irow, column=1).value = "You Lost All~"
        else :
            for i in range(0, len(userRank)) :
                ws.cell(row=irow+i, column=1).value = "{}등 당첨".format(i+1)
                ws.cell(row=irow+i, column=2).value = userRank[i]

        ws.freeze_panes = 'A2'
        wb.save(myExcelPATH+myExcelFile_1)
        wb.close()
        return 0

    # 사용자가 파일에 저장한 번호 조합(6개 숫자)를 읽어오는 메소드
    # ws = wb["Sheet"]  --> get_sheet_by_name 함수는 더이상 사용하지 않음
    # 2018.8.10 현재 미완성 상태.
    # list에 각 row의 튜플 값들을 넣어서 분석할 수 있도록 해야 함.
    # 현재는 단순하게 읽어서 출력하는 정도.
    def loadLottoExcel(self, filePATH) :
        wb = openpyxl.load_workbook(filePATH)
        ws = wb["LottoNum"]

        all_rows = ws.rows
        all_columns = ws.columns

        for row in all_rows :
            for cell in row :
                print(cell.value)

        wb.close()
        return 0

if __name__ == '__main__':
    # 사용자가 임의로 구매한 로또번호를 저장한 파일의 위치
    while 1:
        myLotto = []
        myRank = [0,0,0,0,0]

        newLotto = Lotto()
        for i in range(int(input("How many Lottos do you want? : "))) :
            myLotto.append(newLotto.buyLotto())

        print("\n==========================================")
        print("Win Num: {:2} {:2} {:2} {:2} {:2} {:2}, Bonus Num: {:2}".format(*newLotto.winNum))
        print("==========================================")

        print("\nYour Number is..")

        x = 0
        for i in myLotto :
            print("{:2} {:2} {:2} {:2} {:2} {:2}".format(*i), end=" ")

            myGrade = newLotto.compLottoNum(i)

            if myGrade > 0 :
                myLotto[x].append(myGrade) # 당첨된 경우 맨끝에 순위 추가.
                myRank[myGrade-1] += 1
                print(" ===> {}등 {}".format(myGrade, newLotto.num_match))
            else : print(" ")

            x += 1

        if myRank == [0]*5 : print("\n You Lost All~")
        for i in range(5) :
            if myRank[i] != 0: print('\t{}등 당첨: {}개'.format(i+1,myRank[i]))

        newLotto.saveLottoExcel(myLotto, myRank)
        # newLotto.loadLottoExcel(myExcelPATH)    현재 미완성 메소드

        if input('\nContinue? y/n: ') == 'n': break
